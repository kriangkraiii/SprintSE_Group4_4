#!/usr/bin/env python3
"""Generate UAT Test Case Excel — DriveToSurvive (10 คอลัมน์ ฟอร์มเดิม, เนื้อหา UAT)."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Styles ───────────────────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
LIGHT_BLUE  = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BOLD_FONT   = Font(name="Calibri", bold=True, size=11)
NORMAL_FONT = Font(name="Calibri", size=11)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

COL_WIDTHS = [12, 22, 28, 42, 24, 28, 24, 20, 12, 20]
COL_HEADERS = [
    "Test Case#", "Test Title", "Test Summary", "Test Steps",
    "Test Data", "Expected Result", "Post-condition",
    "Actual Result", "Status", "Notes",
]


def build_sheet(ws, meta, test_cases):
    meta_left_labels  = ["Project Name:", "Module Name:", "Release Version:", ""]
    meta_right_labels = ["Test Designed by:", "Test Designed date:", "Test Executed by:", "Test Execution date:"]
    meta_left_vals  = [meta.get("project",""), meta.get("module",""), meta.get("version",""), ""]
    meta_right_vals = [meta.get("designed_by",""), meta.get("designed_date",""), meta.get("executed_by",""), meta.get("execution_date","")]

    ws.merge_cells("A1:A4")
    ws["A1"] = "DriveToSurvive"
    ws["A1"].font = BOLD_FONT
    ws["A1"].alignment = CENTER
    ws["A1"].fill = LIGHT_BLUE

    for i in range(4):
        row = i + 1
        ws.cell(row=row, column=2, value=meta_left_labels[i]).font = BOLD_FONT
        ws.cell(row=row, column=2).fill = LIGHT_BLUE
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        ws.cell(row=row, column=3, value=meta_left_vals[i]).font = NORMAL_FONT
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
        ws.cell(row=row, column=5, value=meta_right_labels[i]).font = BOLD_FONT
        ws.cell(row=row, column=5).fill = LIGHT_BLUE
        ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=8)
        ws.cell(row=row, column=7, value=meta_right_vals[i]).font = NORMAL_FONT

    for r in range(1, 5):
        for c in range(1, 9):
            ws.cell(row=r, column=c).border = THIN_BORDER

    info_labels = ["Pre-condition", "Dependencies:", "Test Priority"]
    info_values = [meta.get("precondition",""), meta.get("dependencies",""), meta.get("priority","High")]
    for i, (lbl, val) in enumerate(zip(info_labels, info_values)):
        row = 6 + i
        ws.cell(row=row, column=1, value=lbl).font = BOLD_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)
        ws.cell(row=row, column=2, value=val).font = NORMAL_FONT
        ws.cell(row=row, column=2).alignment = WRAP
        for c in range(1, 11):
            ws.cell(row=row, column=c).border = THIN_BORDER

    header_row = 10
    for ci, hdr in enumerate(COL_HEADERS, start=1):
        cell = ws.cell(row=header_row, column=ci, value=hdr)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    for ri, tc in enumerate(test_cases, start=header_row + 1):
        for ci, val in enumerate(tc, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = NORMAL_FONT
            cell.alignment = WRAP
            cell.border = THIN_BORDER

    last = header_row + len(test_cases) + 1
    for ri in range(last, last + 3):
        for ci in range(1, 11):
            ws.cell(row=ri, column=ci).border = THIN_BORDER

    for ci, w in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w


# =====================================================================
#  US#1 — System Log (Compliance) — UAT
# =====================================================================
US1_META = dict(
    project="DriveToSurviveWebApp",
    module="Admin - System Log",
    version="1.0.0",
    designed_by="ทีม QA",
    designed_date="2026-02-16",
    executed_by="",
    execution_date="",
    precondition="เข้าสู่ระบบด้วยบัญชี Admin แล้ว",
    dependencies="Active Admin Account",
    priority="High",
)

US1_CASES = [
    [
        "TC-LOG-001",
        "Open System Log Page",
        "ผู้ดูแลระบบสามารถเปิดหน้า System Log ได้",
        "1. เปิดหน้า Admin Dashboard\n2. คลิกเมนู \"System Logs\" ที่แถบด้านข้าง",
        "Logged in Admin",
        "แสดงหน้า System Log\nพร้อมตารางรายการ Log ที่มี userId, action, ipAddress, createdAt",
        "แสดงหน้า System Log ถูกต้อง",
        "", "", "พ.ร.บ.คอมพิวเตอร์ ม.26"
    ],
    [
        "TC-LOG-002",
        "Filter Log by Date Range",
        "ผู้ดูแลระบบสามารถกรอง Log ตามช่วงวันที่ได้",
        "1. เปิดหน้า System Log\n2. เลือกวันที่เริ่มต้น = \"01/01/2026\"\n3. เลือกวันที่สิ้นสุด = \"31/01/2026\"\n4. กดปุ่ม \"ค้นหา\"",
        "วันที่เริ่ม: 01/01/2026\nวันที่สิ้นสุด: 31/01/2026",
        "แสดงเฉพาะ Log ที่อยู่ในช่วงวันที่ 1-31 ม.ค. 2026",
        "ตารางแสดงผลลัพธ์ที่กรองแล้ว",
        "", "", ""
    ],
    [
        "TC-LOG-003",
        "Filter Log by User ID",
        "ผู้ดูแลระบบสามารถกรอง Log ตาม userId ได้",
        "1. เปิดหน้า System Log\n2. กรอก userId ในช่องค้นหา\n3. กดปุ่ม \"ค้นหา\"",
        "userId ของผู้ใช้ที่ต้องการ",
        "แสดงเฉพาะ Log ของ userId ที่ระบุ",
        "ตารางแสดงผลลัพธ์ที่กรองแล้ว",
        "", "", ""
    ],
    [
        "TC-LOG-004",
        "Filter Log by Action",
        "ผู้ดูแลระบบสามารถกรอง Log ตามประเภท action ได้",
        "1. เปิดหน้า System Log\n2. เลือก action = \"LOGIN\" จาก dropdown\n3. กดปุ่ม \"ค้นหา\"",
        "action = LOGIN",
        "แสดงเฉพาะ Log ที่มี action = LOGIN",
        "ตารางแสดงผลลัพธ์ที่กรองแล้ว",
        "", "", ""
    ],
    [
        "TC-LOG-005",
        "Filter Log by IP Address",
        "ผู้ดูแลระบบสามารถกรอง Log ตาม IP address ได้",
        "1. เปิดหน้า System Log\n2. กรอก IP address = \"192.168\" ในช่องค้นหา\n3. กดปุ่ม \"ค้นหา\"",
        "ipAddress = 192.168",
        "แสดงเฉพาะ Log ที่มี IP address ตรงกับที่ค้นหา",
        "ตารางแสดงผลลัพธ์ที่กรองแล้ว",
        "", "", ""
    ],
    [
        "TC-LOG-006",
        "Pagination Log",
        "ผู้ดูแลระบบสามารถเปลี่ยนหน้าดู Log ได้",
        "1. เปิดหน้า System Log\n2. ดูตารางหน้าแรก\n3. กดปุ่ม \"หน้าถัดไป\" หรือเลือกหมายเลขหน้า",
        "Logged in Admin",
        "แสดง Log หน้าถัดไป\nPagination อัปเดตถูกต้อง",
        "แสดงข้อมูลหน้าที่เลือก",
        "", "", ""
    ],
    [
        "TC-LOG-007",
        "No Edit Button on Log",
        "ระบบไม่มีปุ่มแก้ไข Log (Immutable ตามกฎหมาย)",
        "1. เปิดหน้า System Log\n2. ตรวจสอบทุกแถวในตาราง\n3. มองหาปุ่ม \"แก้ไข\" หรือ \"Edit\"",
        "N/A",
        "ไม่มีปุ่มแก้ไข Log ปรากฏในทุกแถว",
        "Log ไม่สามารถแก้ไขได้",
        "", "", "Immutability"
    ],
    [
        "TC-LOG-008",
        "No Delete Button on Log",
        "ระบบไม่มีปุ่มลบ Log (เก็บ ≥90 วัน ตามกฎหมาย)",
        "1. เปิดหน้า System Log\n2. ตรวจสอบทุกแถวในตาราง\n3. มองหาปุ่ม \"ลบ\" หรือ \"Delete\"",
        "N/A",
        "ไม่มีปุ่มลบ Log ปรากฏในทุกแถว",
        "Log ไม่สามารถลบได้",
        "", "", "พ.ร.บ.คอมพิวเตอร์ ม.26 — เก็บ ≥90 วัน"
    ],
    [
        "TC-LOG-009",
        "Non-Admin Cannot Access Log",
        "ผู้ใช้ทั่วไปไม่สามารถเข้าหน้า System Log ได้",
        "1. เข้าสู่ระบบด้วยบัญชี Passenger\n2. พิมพ์ URL หน้า System Log ในเบราว์เซอร์โดยตรง",
        "Logged in Passenger",
        "ถูก redirect ไปหน้าอื่น\nหรือแสดงข้อความ \"ไม่มีสิทธิ์เข้าถึง\"",
        "ผู้ใช้ทั่วไปไม่เห็นข้อมูล Log",
        "", "", "RBAC"
    ],
    [
        "TC-LOG-010",
        "Log Retained After User Deletion",
        "เมื่อผู้ใช้ถูกลบ Log ของผู้ใช้ยังคงอยู่ในระบบ",
        "1. Admin ลบบัญชีผู้ใช้ A จากหน้าจัดการผู้ใช้\n2. เปิดหน้า System Log\n3. กรอก userId ของผู้ใช้ A ในช่องค้นหา\n4. กดปุ่ม \"ค้นหา\"",
        "userId ของผู้ใช้ที่ถูกลบ",
        "Log ของผู้ใช้ A ยังคงแสดงอยู่ในตาราง",
        "SystemLog ถูกเก็บรักษาไว้\nข้อมูลส่วนบุคคลของ User ถูก anonymize",
        "", "", "PDPA ม.33 vs พ.ร.บ.คอม ม.26"
    ],
]

# =====================================================================
#  US#3 — Blacklist — UAT
# =====================================================================
US3_META = dict(
    project="DriveToSurviveWebApp",
    module="Admin - Blacklist Management",
    version="1.0.0",
    designed_by="ทีม QA",
    designed_date="2026-02-16",
    executed_by="",
    execution_date="",
    precondition="เข้าสู่ระบบด้วยบัญชี Admin แล้ว",
    dependencies="Active Admin Account, มีข้อมูลผู้ใช้ในระบบ",
    priority="High",
)

US3_CASES = [
    [
        "TC-BL-001",
        "Open Blacklist Page",
        "ผู้ดูแลระบบสามารถเปิดหน้าจัดการ Blacklist ได้",
        "1. เปิดหน้า Admin Dashboard\n2. คลิกเมนู \"Blacklist\" ที่แถบด้านข้าง",
        "Logged in Admin",
        "แสดงหน้า Blacklist\nพร้อมตารางรายชื่อที่ถูก blacklist",
        "แสดงหน้า Blacklist ถูกต้อง",
        "", "", ""
    ],
    [
        "TC-BL-002",
        "Add to Blacklist",
        "ผู้ดูแลระบบสามารถเพิ่มเลขบัตรประชาชนเข้า Blacklist ได้",
        "1. เปิดหน้า Blacklist\n2. กดปุ่ม \"เพิ่ม Blacklist\"\n3. กรอกเลขบัตรประชาชน = \"1234567890123\"\n4. กรอกเหตุผล = \"ฉ้อโกง\"\n5. กดปุ่ม \"ยืนยัน\"",
        'เลขบัตร: 1234567890123\nเหตุผล: "ฉ้อโกง"',
        "เพิ่มสำเร็จ\nแสดงรายการใหม่ในตาราง",
        "เพิ่ม Blacklist สำเร็จ ถูกต้อง",
        "", "", "PDPA ม.22 — เก็บเป็น SHA-256 hash"
    ],
    [
        "TC-BL-003",
        "Add Duplicate Blacklist",
        "ระบบป้องกันการเพิ่มเลขบัตรเดียวกันซ้ำ",
        "1. เปิดหน้า Blacklist\n2. กดปุ่ม \"เพิ่ม Blacklist\"\n3. กรอกเลขบัตร = \"1234567890123\" (เพิ่มไปแล้ว)\n4. กดปุ่ม \"ยืนยัน\"",
        'เลขบัตร: 1234567890123\n(เพิ่มไปแล้ว)',
        "แสดงข้อความ \"เลขบัตรนี้อยู่ใน Blacklist แล้ว\"",
        "ไม่มี entry ซ้ำในระบบ",
        "", "", ""
    ],
    [
        "TC-BL-004",
        "Remove from Blacklist",
        "ผู้ดูแลระบบสามารถลบรายชื่อออกจาก Blacklist ได้",
        "1. เปิดหน้า Blacklist\n2. กดปุ่ม \"ลบ\" ที่รายการที่ต้องการ\n3. กดปุ่ม \"ยืนยันการลบ\"",
        "Blacklist entry ที่มีอยู่",
        "ลบสำเร็จ\nรายการหายจากตาราง",
        "ลบ Blacklist สำเร็จ ถูกต้อง",
        "", "", ""
    ],
    [
        "TC-BL-005",
        "Blacklisted Passenger Cannot Register",
        "ผู้ใช้ที่เลขบัตรติด blacklist ไม่สามารถสมัครบัญชีใหม่ (Passenger) ได้",
        "1. Admin เพิ่มเลขบัตร 1234567890123 เข้า blacklist\n2. Logout\n3. เปิดหน้าสมัครสมาชิก\n4. กรอกข้อมูลครบถ้วน เลขบัตร = \"1234567890123\"\n5. กดปุ่ม \"สมัคร\"",
        'เลขบัตร: 1234567890123\nemail: new@test.com\nusername: newuser',
        "แสดงข้อความ \"เลขบัตรนี้ถูกระงับ ไม่สามารถสมัครได้\"",
        "ไม่มีบัญชีใหม่ถูกสร้าง",
        "", "", "ป้องกัน re-apply ด้วยบัญชีใหม่"
    ],
    [
        "TC-BL-006",
        "Blacklisted Driver Cannot Register",
        "ผู้ใช้ที่เลขบัตรติด blacklist ไม่สามารถสมัครเป็น Driver ได้",
        "1. Admin เพิ่มเลขบัตร 1234567890123 เข้า blacklist\n2. Logout\n3. เปิดหน้าสมัครสมาชิก\n4. กรอกข้อมูลครบถ้วน เลขบัตร = \"1234567890123\" เลือก role = Driver\n5. กดปุ่ม \"สมัคร\"",
        'เลขบัตร: 1234567890123\nrole: Driver',
        "แสดงข้อความ \"เลขบัตรนี้ถูกระงับ ไม่สามารถสมัครได้\"",
        "ไม่มีบัญชีใหม่ถูกสร้าง",
        "", "", "ครอบคลุมทั้ง Driver และ Passenger"
    ],
    [
        "TC-BL-007",
        "Clean User Can Register",
        "ผู้ใช้ที่เลขบัตรไม่ติด blacklist สามารถสมัครได้ปกติ",
        "1. เปิดหน้าสมัครสมาชิก\n2. กรอกข้อมูลครบถ้วน เลขบัตร = \"9999999999999\"\n3. อัปโหลดรูปบัตร + selfie\n4. กดปุ่ม \"สมัคร\"",
        'เลขบัตร: 9999999999999\n(ไม่ติด blacklist)',
        "สมัครสำเร็จ\nredirect ไปหน้า login",
        "สมัครสำเร็จ ถูกต้อง",
        "", "", ""
    ],
    [
        "TC-BL-008",
        "Filter Blacklist by Date",
        "ผู้ดูแลระบบสามารถกรอง Blacklist ตามช่วงวันที่ได้",
        "1. เปิดหน้า Blacklist\n2. เลือกวันที่เริ่มต้น = \"01/01/2026\"\n3. เลือกวันที่สิ้นสุด = \"31/12/2026\"\n4. กดปุ่ม \"ค้นหา\"",
        "วันที่เริ่ม: 01/01/2026\nวันที่สิ้นสุด: 31/12/2026",
        "แสดงเฉพาะรายการที่อยู่ในช่วงวันที่ที่เลือก",
        "ตารางแสดงผลลัพธ์ที่กรองแล้ว",
        "", "", ""
    ],
    [
        "TC-BL-009",
        "Non-Admin Cannot Access Blacklist",
        "ผู้ใช้ทั่วไปไม่สามารถเข้าหน้า Blacklist ได้",
        "1. เข้าสู่ระบบด้วยบัญชี Passenger\n2. พิมพ์ URL หน้า Blacklist ในเบราว์เซอร์โดยตรง",
        "Logged in Passenger",
        "ถูก redirect ไปหน้าอื่น\nหรือแสดงข้อความ \"ไม่มีสิทธิ์เข้าถึง\"",
        "ผู้ใช้ทั่วไปไม่เห็นข้อมูล Blacklist",
        "", "", "RBAC"
    ],
    [
        "TC-BL-010",
        "Verify Hash Storage",
        "ระบบเก็บเฉพาะ SHA-256 hash ไม่เก็บเลข 13 หลักดิบ",
        "1. Admin เพิ่มเลขบัตรเข้า blacklist\n2. เปิดหน้า Blacklist\n3. ตรวจสอบคอลัมน์ในตาราง — ไม่แสดงเลขบัตรดิบ",
        "ตรวจสอบหน้า Blacklist",
        "ไม่มีเลขบัตรประชาชน 13 หลักแสดงในตาราง\nแสดงเฉพาะ hash หรือ masked value",
        "ความเป็นส่วนตัวได้รับการปกป้อง",
        "", "", "PDPA ม.22"
    ],
]


def build_report_sheet(ws, sheets_info):
    """
    สร้างหน้า Test Report Summary (Dashboard)
    sheets_info: list of (sheet_title, meta, cases)
    """
    ws.title = "Test Report"
    
    # 1. Header Project Info
    ws.merge_cells("A1:E1")
    ws["A1"] = "Test Execution Report"
    ws["A1"].font = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
    ws["A1"].alignment = CENTER
    ws["A1"].fill = HEADER_FILL

    # Meta Info Table
    meta_keys = ["Project Name:", "Version:", "Test Cycle:", "Date:"]
    # ใช้ Meta ของ US แรกเป็นตัวแทน project info
    first_meta = sheets_info[0][1] if sheets_info else {}
    meta_vals = [
        first_meta.get("project", "DriveToSurviveWebApp"),
        first_meta.get("version", "1.0.0"),
        "UAT-Cycle-1",
        first_meta.get("execution_date", "2026-02-17")
    ]

    for i, (k, v) in enumerate(zip(meta_keys, meta_vals)):
        row = i + 3
        ws.cell(row=row, column=1, value=k).font = BOLD_FONT
        ws.cell(row=row, column=2, value=v).font = NORMAL_FONT
        ws.cell(row=row, column=1).fill = LIGHT_BLUE
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2).border = THIN_BORDER

    # 2. Summary Table Header
    summary_headers = ["Module / Feature", "Total", "Pass", "Fail", "Blocked", "Not Run", "Progress %"]
    header_row = 8
    
    for col_idx, header in enumerate(summary_headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # 3. Fill Data using Formulas
    # Column I in Test Case sheets is "Status" (Index 9)
    # Status Values expected: "Pass", "Fail", "Blocked", "" (Empty = Not Run)
    
    start_data_row = 9
    for i, (sheet_name, meta, cases) in enumerate(sheets_info):
        row = start_data_row + i
        # Safe sheet name for formula (wrap in single quotes)
        safe_sheet_name = f"'{sheet_name}'"
        
        # A: Module Name
        ws.cell(row=row, column=1, value=sheet_name).border = THIN_BORDER
        
        # B: Total Cases (Count all rows in test case sheet)
        # =COUNTA('SheetName'!A:A) - Header Rows (approx 10)
        # Or just hardcode len(cases) for static report, 
        # but formula is better for dynamic updates. 
        # Let's use COUNTIF on ID column, assuming ID starts with "TC-"
        ws.cell(row=row, column=2, value=f'=COUNTIF({safe_sheet_name}!A:A, "TC-*")').border = THIN_BORDER
        
        # C: Pass
        ws.cell(row=row, column=3, value=f'=COUNTIF({safe_sheet_name}!I:I, "Pass")').border = THIN_BORDER
        
        # D: Fail
        ws.cell(row=row, column=4, value=f'=COUNTIF({safe_sheet_name}!I:I, "Fail")').border = THIN_BORDER
        
        # E: Blocked
        ws.cell(row=row, column=5, value=f'=COUNTIF({safe_sheet_name}!I:I, "Blocked")').border = THIN_BORDER
        
        # F: Not Run (Total - (Pass + Fail + Blocked))
        # Cell references: B_row - SUM(C_row:E_row)
        ws.cell(row=row, column=6, value=f'=B{row}-SUM(C{row}:E{row})').border = THIN_BORDER
        
        # G: Progress % ((Pass+Fail+Blocked)/Total)
        ws.cell(row=row, column=7, value=f'=IF(B{row}>0, SUM(C{row}:E{row})/B{row}, 0)').border = THIN_BORDER
        ws.cell(row=row, column=7).number_format = '0%'

    # Total Row
    total_row = start_data_row + len(sheets_info)
    ws.cell(row=total_row, column=1, value="GRAND TOTAL").font = BOLD_FONT
    ws.cell(row=total_row, column=1).border = THIN_BORDER
    
    for col_char in ['B', 'C', 'D', 'E', 'F']:
        col_idx = ord(col_char) - 64
        ws.cell(row=total_row, column=col_idx, value=f'=SUM({col_char}{start_data_row}:{col_char}{total_row-1})').font = BOLD_FONT
        ws.cell(row=total_row, column=col_idx).border = THIN_BORDER

    # Total Progress
    ws.cell(row=total_row, column=7, value=f'=IF(B{total_row}>0, SUM(C{total_row}:E{total_row})/B{total_row}, 0)').font = BOLD_FONT
    ws.cell(row=total_row, column=7).border = THIN_BORDER
    ws.cell(row=total_row, column=7).number_format = '0%'

    # Adjust Widths
    ws.column_dimensions['A'].width = 35
    for col in ['B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col].width = 15



# =====================================================================
#  US#4 — Automated Verification — UAT
# =====================================================================
US4_META = dict(
    project="DriveToSurviveWebApp",
    module="Automated User Verification",
    version="1.0.0",
    designed_by="ทีม QA",
    designed_date="2026-02-16",
    executed_by="",
    execution_date="",
    precondition="มีบัญชีผู้ใช้และเข้าสู่ระบบแล้ว",
    dependencies="Active User Account, Cloudinary, iApp OCR Service",
    priority="High",
)

US4_CASES = [
    [
        "TC-VER-001",
        "Register with Full Docs — Auto Verified",
        "ผู้ใช้สมัครสมาชิกพร้อมเอกสารครบ ระบบยืนยันตัวตนอัตโนมัติ",
        "1. เปิดหน้าสมัครสมาชิก\n2. กรอกข้อมูลครบถ้วน (ชื่อ, นามสกุล, email, เบอร์โทร, เลขบัตร)\n3. อัปโหลดรูปบัตรประชาชน\n4. อัปโหลดรูป selfie\n5. กดปุ่ม \"สมัคร\"",
        "ข้อมูลครบถ้วน\nรูปบัตร + รูป selfie",
        "สมัครสำเร็จ\nสถานะยืนยันตัวตน = ยืนยันแล้ว (ทันที)\nไม่ต้องรอ Admin อนุมัติ",
        "ผู้ใช้เข้าถึงฟีเจอร์ที่ต้องยืนยันตัวตนได้เลย",
        "", "", "Auto-verify ตอนสมัคร"
    ],
    [
        "TC-VER-002",
        "Register without Docs — Not Verified",
        "ผู้ใช้สมัครสมาชิกโดยไม่อัปโหลดเอกสาร ไม่ถูกยืนยันอัตโนมัติ",
        "1. เปิดหน้าสมัครสมาชิก\n2. กรอกข้อมูลครบถ้วน\n3. ไม่อัปโหลดรูปบัตรประชาชน\n4. กดปุ่ม \"สมัคร\"",
        "ข้อมูลครบ แต่ไม่มีรูปบัตร/selfie",
        "สมัครสำเร็จ\nสถานะยืนยันตัวตน = ยังไม่ยืนยัน",
        "ผู้ใช้ต้องส่งเอกสารเพิ่มเติมภายหลัง",
        "", "", ""
    ],
    [
        "TC-VER-003",
        "Submit Driver License — Auto Approved",
        "ผู้ใช้ส่งใบขับขี่ ระบบอนุมัติอัตโนมัติโดยไม่ต้องรอ Admin",
        "1. เข้าสู่ระบบ\n2. เปิดหน้า \"ยืนยันตัวตนคนขับ\"\n3. กรอกเลขใบขับขี่ = \"12345678\"\n4. กรอกชื่อ-นามสกุลบนใบขับขี่\n5. เลือกประเภทใบขับขี่\n6. เลือกวันออก/วันหมดอายุ\n7. อัปโหลดรูปใบขับขี่\n8. อัปโหลดรูป selfie\n9. กดปุ่ม \"ส่งยืนยัน\"",
        'เลขใบขับขี่: 12345678\nรูปใบขับขี่ + รูป selfie',
        "ส่งสำเร็จ\nสถานะ = \"อนุมัติแล้ว\"\nผู้ใช้ได้รับการแจ้งเตือน",
        "ส่งยืนยันสำเร็จ ถูกต้อง\nผู้ใช้สามารถสร้างเส้นทางได้",
        "", "", "ฟีเจอร์หลัก Auto-verification"
    ],
    [
        "TC-VER-004",
        "Submit without License Photo — Rejected",
        "ผู้ใช้ส่งยืนยันโดยไม่มีรูปใบขับขี่ ระบบปฏิเสธ",
        "1. เข้าสู่ระบบ\n2. เปิดหน้า \"ยืนยันตัวตนคนขับ\"\n3. กรอกข้อมูลครบ\n4. อัปโหลดเฉพาะรูป selfie (ไม่อัปโหลดรูปใบขับขี่)\n5. กดปุ่ม \"ส่งยืนยัน\"",
        "มีเฉพาะรูป selfie",
        "แสดงข้อความ \"กรุณาอัปโหลดรูปใบขับขี่และรูป selfie\"",
        "ไม่มี verification record ถูกสร้าง",
        "", "", ""
    ],
    [
        "TC-VER-005",
        "Submit without Selfie — Rejected",
        "ผู้ใช้ส่งยืนยันโดยไม่มีรูป selfie ระบบปฏิเสธ",
        "1. เข้าสู่ระบบ\n2. เปิดหน้า \"ยืนยันตัวตนคนขับ\"\n3. กรอกข้อมูลครบ\n4. อัปโหลดเฉพาะรูปใบขับขี่ (ไม่อัปโหลด selfie)\n5. กดปุ่ม \"ส่งยืนยัน\"",
        "มีเฉพาะรูปใบขับขี่",
        "แสดงข้อความ \"กรุณาอัปโหลดรูปใบขับขี่และรูป selfie\"",
        "ไม่มี verification record ถูกสร้าง",
        "", "", ""
    ],
    [
        "TC-VER-006",
        "OCR Auto Verification",
        "ผู้ใช้ส่งยืนยันผ่าน OCR ระบบอ่านข้อมูลและอนุมัติอัตโนมัติ",
        "1. เข้าสู่ระบบ\n2. เปิดหน้า \"ยืนยันตัวตนคนขับ\"\n3. เลือกโหมด \"OCR อัตโนมัติ\"\n4. อัปโหลดรูปใบขับขี่\n5. อัปโหลดรูป selfie\n6. กดปุ่ม \"ส่งยืนยัน\"",
        "รูปใบขับขี่ + รูป selfie",
        "ระบบอ่านข้อมูลจากรูปอัตโนมัติ\nสถานะ = \"อนุมัติแล้ว\"\nข้อมูล OCR ถูกบันทึก",
        "ส่งยืนยัน OCR สำเร็จ ถูกต้อง",
        "", "", "iApp OCR"
    ],
    [
        "TC-VER-007",
        "Re-submit Updates Existing Record",
        "ผู้ใช้ส่งยืนยันซ้ำ ระบบอัปเดต record เดิมแทนการสร้างซ้ำ",
        "1. เข้าสู่ระบบ (ผู้ใช้ที่เคยส่งยืนยันแล้ว)\n2. เปิดหน้า \"ยืนยันตัวตนคนขับ\"\n3. อัปโหลดรูปใหม่\n4. กดปุ่ม \"ส่งยืนยัน\"",
        "ผู้ใช้คนเดิม, รูปใหม่",
        "อัปเดตสำเร็จ\nRecord เดิมถูกอัปเดต (ไม่สร้างซ้ำ)\nสถานะ = \"อนุมัติแล้ว\"",
        "มี verification record เดียวต่อผู้ใช้",
        "", "", ""
    ],
    [
        "TC-VER-008",
        "Notification After Auto Approval",
        "ผู้ใช้ได้รับการแจ้งเตือนหลังอนุมัติอัตโนมัติ",
        "1. ส่งยืนยันตัวตนคนขับสำเร็จ\n2. คลิกไอคอนแจ้งเตือน (กระดิ่ง)\n3. ตรวจสอบรายการแจ้งเตือน",
        "Logged in User",
        "มีการแจ้งเตือนใหม่\nหัวข้อ: \"ยืนยันตัวตนคนขับสำเร็จ\"",
        "ผู้ใช้มีการแจ้งเตือนที่ยังไม่ได้อ่าน",
        "", "", ""
    ],
    [
        "TC-VER-009",
        "Admin Override — Reject Verification",
        "Admin สามารถปฏิเสธการยืนยันที่อนุมัติอัตโนมัติแล้วได้ (กรณีพิเศษ)",
        "1. เข้าสู่ระบบด้วย Admin\n2. เปิดหน้า \"จัดการการยืนยันตัวตน\"\n3. เลือกรายการที่สถานะ \"อนุมัติแล้ว\"\n4. กดปุ่ม \"เปลี่ยนสถานะ\"\n5. เลือก \"ปฏิเสธ\"\n6. กดปุ่ม \"ยืนยัน\"",
        "Verification ที่สถานะ APPROVED",
        "สถานะเปลี่ยนเป็น \"ปฏิเสธ\"\nเส้นทางที่สถานะ AVAILABLE ถูกยกเลิก",
        "คนขับไม่สามารถสร้างเส้นทางใหม่ได้",
        "", "", "Admin จัดการกรณีพิเศษ"
    ],
    [
        "TC-VER-010",
        "Admin Manual Approve",
        "Admin สามารถอนุมัติการยืนยันด้วยตนเองได้",
        "1. เข้าสู่ระบบด้วย Admin\n2. เปิดหน้า \"จัดการการยืนยันตัวตน\"\n3. เลือกรายการที่ต้องการ\n4. กดปุ่ม \"เปลี่ยนสถานะ\"\n5. เลือก \"อนุมัติ\"\n6. กดปุ่ม \"ยืนยัน\"",
        "Verification record",
        "สถานะเปลี่ยนเป็น \"อนุมัติแล้ว\"\nผู้ใช้ isVerified = true",
        "ผู้ใช้สามารถสร้างเส้นทางได้",
        "", "", ""
    ],
    [
        "TC-VER-011",
        "Verified Driver Can Create Route",
        "คนขับที่ยืนยันแล้วสามารถสร้างเส้นทางได้",
        "1. เข้าสู่ระบบด้วยคนขับที่ยืนยันแล้ว\n2. เปิดหน้า \"สร้างเส้นทาง\"\n3. กรอกข้อมูลเส้นทาง\n4. กดปุ่ม \"สร้าง\"",
        "Logged in Verified Driver",
        "สร้างเส้นทางสำเร็จ",
        "มีเส้นทางใหม่ในระบบ",
        "", "", ""
    ],
    [
        "TC-VER-012",
        "Unverified User Cannot Create Route",
        "ผู้ใช้ที่ยังไม่ยืนยันไม่สามารถสร้างเส้นทางได้",
        "1. เข้าสู่ระบบด้วยผู้ใช้ที่ยังไม่ยืนยัน\n2. พยายามเปิดหน้า \"สร้างเส้นทาง\"",
        "Logged in Unverified User",
        "ไม่สามารถสร้างเส้นทางได้\nแสดงข้อความ \"กรุณายืนยันตัวตนก่อน\"",
        "ไม่มีเส้นทางใหม่ถูกสร้าง",
        "", "", ""
    ],
]

# =====================================================================
#  US#16 — Account Deletion — UAT
# =====================================================================
US16_META = dict(
    project="DriveToSurviveWebApp",
    module="Profile - Delete Account",
    version="1.0.0",
    designed_by="ทีม QA",
    designed_date="2026-02-16",
    executed_by="",
    execution_date="",
    precondition="เข้าสู่ระบบแล้ว",
    dependencies="Active User Account",
    priority="High",
)

US16_CASES = [
    [
        "TC-DEL-001",
        "Open Delete Account Modal",
        "ผู้ใช้สามารถเปิดหน้าลบบัญชีได้",
        "1. เปิดหน้า Profile\n2. เลื่อนลงมาที่ส่วน \"ลบบัญชีผู้ใช้\"\n3. กดปุ่ม \"ลบบัญชีของฉัน\"",
        "Logged in User",
        "แสดง Modal ยืนยันการลบบัญชี",
        "แสดง Modal ถูกต้อง",
        "", "", ""
    ],
    [
        "TC-DEL-002",
        "Confirm Delete Account",
        "ผู้ใช้สามารถลบบัญชีโดยพิมพ์ DELETE ยืนยัน",
        "1. พิมพ์ \"DELETE\" ในช่อง input\n2. กดปุ่มยืนยัน",
        "DELETE",
        "บัญชีถูกลบ และ redirect ไปหน้า login",
        "ลบบัญชีสำเร็จ ถูกต้อง",
        "", "", "PDPA ม.33"
    ],
    [
        "TC-DEL-003",
        "Cancel Delete",
        "ผู้ใช้สามารถยกเลิกการลบบัญชีได้",
        "1. กดปุ่ม \"ยกเลิก\" ใน Modal",
        "N/A",
        "Modal ปิด\nและยังอยู่ที่หน้า Profile",
        "Modal ปิดถูกต้อง",
        "", "", ""
    ],
    [
        "TC-DEL-004",
        "Wrong Confirm Text",
        "ระบบตรวจสอบข้อความยืนยัน ถ้าพิมพ์ไม่ตรงจะไม่สามารถลบได้",
        "1. พิมพ์ \"delete\" (ตัวเล็ก) ในช่อง input",
        "delete (lowercase)",
        "ปุ่มยืนยันถูก disable",
        "ปุ่มถูก disable ถูกต้อง",
        "", "", ""
    ],
    [
        "TC-DEL-005",
        "Deleted User Cannot Login",
        "ผู้ใช้ที่ลบบัญชีแล้วไม่สามารถเข้าสู่ระบบได้",
        "1. ลบบัญชีสำเร็จ (จาก TC-DEL-002)\n2. เปิดหน้า Login\n3. กรอก email และ password เดิม\n4. กดปุ่ม \"เข้าสู่ระบบ\"",
        "email และ password เดิม",
        "เข้าสู่ระบบไม่ได้\nแสดงข้อความ \"ไม่พบบัญชีผู้ใช้\"",
        "ไม่มี session ถูกสร้าง",
        "", "", ""
    ],
    [
        "TC-DEL-006",
        "PII Anonymized After Deletion",
        "ข้อมูลส่วนบุคคลถูก anonymize หลังลบบัญชี",
        "1. Admin เข้าสู่ระบบ\n2. เปิดหน้าจัดการผู้ใช้\n3. ค้นหาผู้ใช้ที่ถูกลบ\n4. ตรวจสอบข้อมูลที่แสดง",
        "userId ของผู้ใช้ที่ถูกลบ",
        "ชื่อ = \"Deleted User\"\nemail = \"deleted_xxx@removed.local\"\nเบอร์โทร = ว่าง\nรูปภาพ = ว่าง",
        "ไม่มีข้อมูลส่วนบุคคลเดิมปรากฏ",
        "", "", "ตรวจสอบ anonymize"
    ],
    [
        "TC-DEL-007",
        "System Log Retained After Deletion",
        "Log ของผู้ใช้ยังคงอยู่หลังลบบัญชี (ตาม พ.ร.บ.คอมพิวเตอร์ ม.26)",
        "1. Admin เข้าสู่ระบบ\n2. เปิดหน้า System Log\n3. กรอก userId ของผู้ใช้ที่ถูกลบ\n4. กดปุ่ม \"ค้นหา\"",
        "userId ของผู้ใช้ที่ถูกลบ",
        "Log ของผู้ใช้ยังคงแสดงอยู่ในตาราง\nจำนวน Log ไม่เปลี่ยนแปลง",
        "SystemLog ถูกเก็บรักษาไว้ ≥90 วัน",
        "", "", "PDPA vs พ.ร.บ.คอม"
    ],
    [
        "TC-DEL-008",
        "Admin Delete User",
        "Admin สามารถลบบัญชีผู้ใช้คนอื่นได้",
        "1. เข้าสู่ระบบด้วย Admin\n2. เปิดหน้าจัดการผู้ใช้\n3. ค้นหาผู้ใช้ที่ต้องการลบ\n4. กดปุ่ม \"ลบ\"\n5. กดปุ่ม \"ยืนยันการลบ\"",
        "userId ของผู้ใช้เป้าหมาย",
        "ลบสำเร็จ\nข้อมูลส่วนบุคคลถูก anonymize\nLog ยังคงอยู่",
        "ลบบัญชีสำเร็จ ถูกต้อง",
        "", "", ""
    ],
    [
        "TC-DEL-009",
        "Delete Already Deleted Account",
        "ลบบัญชีที่ถูกลบแล้วซ้ำ แสดง error",
        "1. Admin เข้าสู่ระบบ\n2. เปิดหน้าจัดการผู้ใช้\n3. ค้นหาผู้ใช้ที่ถูกลบแล้ว\n4. กดปุ่ม \"ลบ\" อีกครั้ง",
        "userId ของผู้ใช้ที่ถูกลบแล้ว",
        "แสดงข้อความ \"บัญชีนี้ถูกลบไปแล้ว\"",
        "ไม่มีการเปลี่ยนแปลงเพิ่มเติม",
        "", "", ""
    ],
    [
        "TC-DEL-010",
        "Non-Admin Cannot Delete Others",
        "ผู้ใช้ทั่วไปไม่สามารถลบบัญชีผู้อื่นได้",
        "1. เข้าสู่ระบบด้วยบัญชี Passenger\n2. พิมพ์ URL หน้าจัดการผู้ใช้ในเบราว์เซอร์โดยตรง",
        "Logged in Passenger",
        "ถูก redirect ไปหน้าอื่น\nหรือแสดงข้อความ \"ไม่มีสิทธิ์เข้าถึง\"",
        "ผู้ใช้เป้าหมายไม่เปลี่ยนแปลง",
        "", "", "RBAC"
    ],
]


# =====================================================================
#  BUILD WORKBOOK
# =====================================================================
wb = openpyxl.Workbook()

sheets = [
    ("US1 - Log ตามกฎหมาย", US1_META, US1_CASES),
    ("US3 - Blacklist", US3_META, US3_CASES),
    ("US4 - ยืนยันตัวตนอัตโนมัติ", US4_META, US4_CASES),
    ("US16 - ลบบัญชีผู้ใช้", US16_META, US16_CASES),
]

# Create Report Sheet first
report_ws = wb.active
build_report_sheet(report_ws, sheets)

for i, (title, meta, cases) in enumerate(sheets):
    ws = wb.create_sheet(title=title)
    build_sheet(ws, meta, cases)

OUTPUT = "/Users/conan/Desktop/ไม่มีชื่อโฟลเดอร์/TestCases_DriveToSurvive.xlsx"
wb.save(OUTPUT)
print(f"✅ UAT Test case document saved to:\n{OUTPUT}")
