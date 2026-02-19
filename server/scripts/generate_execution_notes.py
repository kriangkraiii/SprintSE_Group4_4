
import openpyxl
import os
import re

# Configuration
EXCEL_PATH = '/Users/conan/Downloads/ซอฟเอ็น/ไฟล์เพิ่มเติม/TestCases_DriveToSurviveใกล้เสร็จ.xlsx'
OUTPUT_FILE = '/Users/conan/Downloads/ซอฟเอ็น/SprintSE_Group4_4/full_test_execution_records.txt'

def clean_text(text):
    if not text: return ""
    # Remove leading numbering like "1. ", "2. "
    lines = []
    for line in str(text).split('\n'):
        line = line.strip()
        if not line: continue
        # Remove "1. ", "1)", etc at start
        cleaned = re.sub(r'^\d+[\.\)]\s*', '', line)
        lines.append(f"- {cleaned}")
    return "\n".join(lines)

def format_steps_observed(steps_text):
    """Formats steps for the 'Having -' section"""
    if not steps_text: return "- Performed test steps."
    return clean_text(steps_text)

def format_expected_verified(expected_text):
    """Formats expected results for the 'I was able to verify that' section"""
    if not expected_text: return "- The result matched the expected outcome."
    return clean_text(expected_text)

def generate_records():
    print("Generating Full Test Execution Records...")
    
    if not os.path.exists(EXCEL_PATH):
        print(f"Error: Excel file not found at {EXCEL_PATH}")
        return

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    
    # Sheets to process (include US4 if it has data now)
    sheets = ['US1 - Log ตามกฎหมาย', 'US3 - Blacklist', 'US4 - ยืนยันตัวตนอัตโนมัติ', 'US16 - ลบบัญชีผู้ใช้']
    
    all_records = []
    
    for sheet_name in sheets:
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            print(f"Processing sheet: {sheet_name}")
            
            # Find header row
            header_row_idx = -1
            for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                if row and row[0] == 'Test Case#':
                    header_row_idx = i
                    break
            
            if header_row_idx == -1:
                print(f"  Warning: Header row not found in {sheet_name}")
                continue
                
            # Read data rows
            for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
                test_id = row[0]
                if not test_id or not str(test_id).startswith('TC-'): continue
                
                title = row[1]
                steps = row[3]
                expected_result = row[5]
                
                # Format content
                steps_formatted = format_steps_observed(steps)
                expected_formatted = format_expected_verified(expected_result)
                
                record = f"""
================================================================================
Test ID: {test_id}
Test Title: {title}
--------------------------------------------------------------------------------
Observed Test Result
--------------------
Having –
{steps_formatted}

I was able to verify that:
{expected_formatted}
"""
                all_records.append(record)
    
    # Write to file
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        f.write("\n".join(all_records))
        
    print(f"Successfully generated {len(all_records)} records to: {OUTPUT_FILE}")

if __name__ == "__main__":
    generate_records()
