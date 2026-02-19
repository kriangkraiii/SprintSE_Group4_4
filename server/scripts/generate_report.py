
import openpyxl
import xml.etree.ElementTree as ET
import json
import os
import re

# Configuration
EXCEL_PATH = '/Users/conan/Downloads/ซอฟเอ็น/ไฟล์เพิ่มเติม/TestCases_DriveToSurviveใกล้เสร็จ.xlsx'
XML_PATH = '/Users/conan/Downloads/ซอฟเอ็น/SprintSE_Group4_4/output.xml'
OUTPUT_JSON_PATH = '/Users/conan/Downloads/ซอฟเอ็น/SprintSE_Group4_4/DriveToSurviveWebApp/client/public/test_report.json'

def parse_robot_xml(xml_path):
    """Parses Robot Framework output.xml and returns a dict mapping Test Case ID to Status."""
    if not os.path.exists(xml_path):
        print(f"Warning: XML file not found at {xml_path}")
        return {}

    tree = ET.parse(xml_path)
    root = tree.getroot()
    
    results = {}
    
    # Iterate through all 'test' elements
    for test in root.iter('test'):
        name = test.get('name')
        # Extract Test Case ID (e.g., TC-LOG-001)
        # Assuming ID is the first word
        parts = name.split(' ', 1)
        if len(parts) > 0:
            test_id = parts[0]
            status_elem = test.find('status')
            if status_elem is not None:
                status = status_elem.get('status') # PASS / FAIL
                results[test_id] = status
                
    return results

def parse_excel(excel_path):
    """Parses Excel file and extracts test cases."""
    if not os.path.exists(excel_path):
        print(f"Error: Excel file not found at {excel_path}")
        return []

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    test_cases = []
    
    # Sheets to process
    sheets = ['US1 - Log ตามกฎหมาย', 'US3 - Blacklist', 'US16 - ลบบัญชีผู้ใช้']
    
    for sheet_name in sheets:
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            print(f"Processing sheet: {sheet_name}")
            
            # Find header row
            header_row_idx = -1
            for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                # Check for "Test Case#" in headers
                if row and row[0] == 'Test Case#':
                    header_row_idx = i
                    break
            
            if header_row_idx == -1:
                print(f"  Warning: Header row not found in {sheet_name}")
                continue
                
            # Read data rows
            for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
                test_id = row[0]
                if not test_id: continue # Skip empty rows
                
                # Check if it looks like a valid ID (TC-xxx)
                if not str(test_id).startswith('TC-'): continue
                
                title = row[1]
                steps = row[3]
                expected_result = row[5]
                
                test_cases.append({
                    'id': test_id,
                    'title': title,
                    'steps': steps,
                    'expected_result': expected_result,
                    'sheet': sheet_name
                })
                
    return test_cases

def generate_report():
    print("Generating Test Report...")
    
    # 1. Get Automation Results
    robot_results = parse_robot_xml(XML_PATH)
    print(f"Loaded {len(robot_results)} automated test results.")
    
    # 2. Get Manual Test Cases
    manual_tests = parse_excel(EXCEL_PATH)
    print(f"Loaded {len(manual_tests)} test cases from Excel.")
    
    # 3. Merge Data
    summary = {
        'total': 0,
        'pass': 0,
        'fail': 0,
        'not_run': 0
    }
    
    details = []
    
    for tc in manual_tests:
        test_id = tc['id']
        status = 'NOT RUN'
        
        # Check if executed by Robot
        if test_id in robot_results:
            status = robot_results[test_id]
        
        # Update summary
        summary['total'] += 1
        if status == 'PASS':
            summary['pass'] += 1
        elif status == 'FAIL':
            summary['fail'] += 1
        else:
            summary['not_run'] += 1
            
        details.append({
            'id': test_id,
            'title': tc['title'],
            'steps': tc['steps'],
            'expected_result': tc['expected_result'],
            'status': status,
            'sheet': tc['sheet']
        })
        
    output_data = {
        'summary': summary,
        'details': details
    }
    
    # 4. Write JSON
    os.makedirs(os.path.dirname(OUTPUT_JSON_PATH), exist_ok=True)
    with open(OUTPUT_JSON_PATH, 'w', encoding='utf-8') as f:
        json.dump(output_data, f, ensure_ascii=False, indent=2)
        
    print(f"Report generated successfully at: {OUTPUT_JSON_PATH}")
    print(f"Summary: {summary}")

if __name__ == "__main__":
    generate_report()
